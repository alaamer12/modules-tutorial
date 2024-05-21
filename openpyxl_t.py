import os


def create():
    from openpyxl import Workbook

    # Create a new workbook
    wb = Workbook()

    ws = wb.active  # Get the active worksheet
    # noinspection PyRedeclaration
    # ws = wb["Sheet1"]  # Get a worksheet by name
    # # noinspection PyRedeclaration
    # ws = wb.worksheets[0]  # Get a worksheet by index

    # Write some data to the worksheet
    ws["A1"] = "Hello"
    ws["B1"] = "World"

    # Read some data from the worksheet
    print(ws["A1"].value)

    # Save the workbook
    wb.save("excels/example.xlsx")


def load():
    from openpyxl import load_workbook

    # Load the workbook
    wb = load_workbook("excels/example.xlsx")

    ws = wb.active  # Get the active worksheet

    # Get some data from the worksheet
    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
        for cell in row:
            print(cell.value)

    # Merge cells
    ws.merge_cells('A1:D1')


def adding_chart():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference

    wb = Workbook()
    ws = wb.active

    data = [
        ['Apples', 100],
        ['Bananas', 150],
        ['Cherries', 200],
    ]

    for row in data:
        ws.append(row)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Fruit Distribution"
    chart.x_axis.title = "Fruit"
    chart.y_axis.title = "Quantity"

    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=4)
    categories = Reference(ws, min_col=1, min_row=2, max_row=4)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    ws.add_chart(chart, "E1")

    wb.save("excels/chart.xlsx")


def conditional_formatting():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import CellIsRule

    wb = Workbook()
    ws = wb.active

    data = [
        [10, 20, 30],
        [40, 50, 60],
        [70, 80, 90]
    ]

    for row in data:
        ws.append(row)

    red_fill = PatternFill(start_color='FF0000', end_color='6e7eff', fill_type='solid')
    rule = CellIsRule(operator='greaterThan', formula=['50'], fill=red_fill)

    ws.conditional_formatting.add('A1:C3', rule)

    wb.save("excels/conditional_formatting.xlsx")


def data_validation():
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active

    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=100)
    ws.add_data_validation(dv)
    dv.add('A1:C3')

    wb.save("excels/data_validation.xlsx")


def main():
    create()
    load()
    adding_chart()
    conditional_formatting()
    data_validation()


if __name__ == "__main__":
    if not os.path.exists("excels"):
        os.mkdir("excels")
    main()
